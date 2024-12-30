"""
Driver list script.
.. moduleauthor::
    Luis Pereira  <luis.pereira@indt.org.br>
    Lahis Almeida <lahis.almeida@indt.org.br>
Usage:
    $ python .\driver_list.py
"""
import wmi
from datetime import datetime
import json
import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font


def driver_check(intel_drivers, system_info):
    file = Workbook()
    sheet = file.active
    sheet.title = "Dataset"
    # Define the "results" directory based on the current file path
    results_dir = os.path.join(os.path.dirname(__file__), "../results")
    os.makedirs(results_dir, exist_ok=True)  # Create the 'results' folder if it doesn't exist
    # Define the header
    header = ["No.", "Drivers", "Installed Version", "Installed Version Date", "Manufacturer",
              "Most Recent Version", "Most Recent Version Date", "Outdated Time"]
    total_columns = len(header)
    merge_range_title = f"A1:{get_column_letter(total_columns)}1"
    merge_range_verification = f"A2:{get_column_letter(total_columns)}2"
    # Fill styles
    title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    verification_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    # Add the merged title row
    sheet.merge_cells(merge_range_title)
    manufacturer = system_info.get("ProductName", "Unknown")
    baseboard = system_info.get("BaseBoard_Product", "Unknown")
    title_text = f"Device Driver Report \"{manufacturer}\" : \"{baseboard}\""
    title_cell = sheet["A1"]
    title_cell.value = title_text
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = title_fill
    sheet.row_dimensions[1].height = 25  # Adjust the height of row 1
    # Add the merged verification date row
    sheet.merge_cells(merge_range_verification)
    current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    verification_text = f"Last Verification: {current_date}"
    verification_cell = sheet["A2"]
    verification_cell.value = verification_text
    verification_cell.alignment = Alignment(horizontal="center", vertical="center")
    verification_cell.font = Font(italic=True, size=12)
    verification_cell.fill = verification_fill
    sheet.row_dimensions[2].height = 20  # Adjust the height of row 2
    # Add the header in row 3
    sheet.append(header)
    # Define border styles
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    for index, driver in enumerate(intel_drivers, start=1):
        sheet.append([index, driver['Name'], driver['Version'], driver['Date'], driver['Manufacturer']])
        # Center content in the current row cells
        for col_num in range(1, len(header) + 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f"{col_letter}{index + 3}"]  # Row index + 3 due to header
            # Set alignment for the cell
            if header[col_num - 1] == "Drivers":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    # Center the header
    for col_num, col_header in enumerate(header, start=1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}3"]  # Center header row
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.fill = header_fill
        sheet.row_dimensions[3].height = 20
        # Set custom width for "Driver"
        if col_header == "No.":
            sheet.column_dimensions[col_letter].width = len(col_header) + 2
        elif col_header == "Drivers":
            sheet.column_dimensions[col_letter].width = max(len(header) + 10, 50)  # Wider width
        else:
            sheet.column_dimensions[col_letter].width = max(len(header) + 5, 25)
    # Define the full Excel file path
    current_date = datetime.now().strftime("%d/%m/%Y")
    excel_file_name = f"Drivers-[{manufacturer}]-[{baseboard}]-[{current_date.replace('/', '-')}].xlsx"
    excel_file_path = os.path.join(results_dir, excel_file_name)
    file.save(excel_file_path)
    os.startfile(excel_file_path)
    print("Excel file successfully generated!!")


def list_total_drivers():
    # Initialize the WMI client
    c = wmi.WMI()

    drivers = []
    # Query installed drivers on the system
    for driver in c.Win32_PnPSignedDriver():
        drivers.append({
            'Name': driver.DeviceName,
            'Manufacturer': driver.Manufacturer,
            'Version': driver.DriverVersion,
            'Class': driver.DeviceClass,
            'Date': driver.DriverDate
        })
    return drivers


def format_wmi_date(wmi_date):
    try:
        # Extract the part of the date before the dot
        formatted_date = datetime.strptime(wmi_date.split('.')[0], "%Y%m%d%H%M%S")
        return formatted_date.strftime("%d/%m/%Y")
    except Exception:
        return "Invalid date"


def list_intel_drivers():
    # Initialize the WMI client
    c = wmi.WMI()
    driver_count = set()
    intel_drivers = []
    # Query installed drivers on the system
    for driver in c.Win32_PnPSignedDriver():

        # Check if the manufacturer is invalid (None, contains "Microsoft", "Standard", or other values)
        # and if the device class is not "PROCESSOR"
        manufacturer = driver.Manufacturer.strip().lower() if driver.Manufacturer else ""
        device_class = driver.DeviceClass.strip().lower() if driver.DeviceClass else None
        if manufacturer and "microsoft" not in manufacturer and \
                "generic" not in manufacturer and \
                "winusb" not in manufacturer and \
                "oracle" not in manufacturer and \
                "standard" not in manufacturer and \
                device_class is not None and \
                "processor" not in device_class:  # Check if the class is not "PROCESSOR"
            driver_count.add(driver.DeviceName)
            intel_drivers.append({
                'Name': driver.DeviceName,
                'Manufacturer': driver.Manufacturer,
                'Version': driver.DriverVersion,
                'Class': driver.DeviceClass,
                'Date': format_wmi_date(driver.DriverDate),
                'HardwareID': driver.HardWareID,
                'DeviceID': driver.DeviceID,
                'CompatID': driver.CompatID
            })
    print("Propriedades disponíveis no objeto 'driver':")
    for prop in driver.properties.keys():
        print(prop)
    system_info = {}

    for system in c.Win32_ComputerSystem():
        system_info['Manufacturer'] = system.Manufacturer
        system_info['Family'] = system.SystemFamily
        system_info['ProductName'] = system.Model
        system_info['SKU'] = system.SystemSKUNumber
        # system_info['NetworkName'] = system.PrimaryOwnerName
        # system_info['InstallDate'] = system.InstallDate
        # system_info['SKU'] = system.SystemSKUNumber
        # system_info['SKU'] = system.SystemSKUNumber
    for chassis in c.Win32_SystemEnclosure():
        # system_info['Chassis_Type'] = chassis.ChassisTypes[0] if chassis.ChassisTypes else "Unknown"
        system_info['SerialNumber'] = chassis.SerialNumber
    for bios in c.Win32_BIOS():
        system_info['BIOSVersion'] = bios.Name
    for processor in c.Win32_Processor():
        system_info['Processor'] = processor.Name
    for baseboard in c.Win32_BaseBoard():
        # windows_info['BaseBoard_Manufacturer'] = baseboard.Manufacturer
        system_info['BaseBoard_Product'] = baseboard.Product
        # windows_info['BaseBoard_SerialNumber'] = baseboard.SerialNumber
    driver_check(intel_drivers, system_info)
    windows_info = {}
    for os in c.Win32_OperatingSystem():
        windows_info['NetworkName'] = os.RegisteredUser
        windows_info['OSDescription'] = os.Caption
        windows_info['InstallDate'] = os.InstallDate
        windows_info['BuildNumber'] = os.BuildNumber
        windows_info['Architecturer'] = os.OSArchitecture
        # windows_info['LockdownPolicy'] = os.DataExecutionPrevention_SupportPolicy
        windows_info['Version'] = os.Version
        windows_info['BaseLanguage'] = os.MUILanguages
        windows_info['MUILanguages'] = os.MUILanguages
    result = {
        'SystemConfig': system_info,
        'WindowsConfig': windows_info,
        'Drivers': intel_drivers,
    }
    return json.dumps(result, ensure_ascii=False, indent=4), len(driver_count)


def json_to_csv(csv_file_name):
    # Sets the full path to the file inside the “results” folder
    results_dir = os.path.join(os.path.dirname(__file__), "../results")
    os.makedirs(results_dir, exist_ok=True)  # Create the 'results' folder if it doesn't exist

    # Full path of the CSV file
    csv_file_path = os.path.join(results_dir, f"{csv_file_name}.csv")
    # Converting JSON to CSV
    with open(csv_file_path, mode="w", newline="") as file:
        writer = csv.writer(file)

        # Write the JSON keys as headers
        writer.writerow(json_result.keys())

        # Write the values from the JSON in the next line
        writer.writerow(json_result.values())
    print(f"JSON converted to CSV in the file {csv_file_path}")


if __name__ == '__main__':
    # Get the drivers found and the quantity of drivers
    drivers_found, quantity_drivers = list_intel_drivers()
    # Parse the JSON result
    json_result = json.loads(drivers_found)

    # 3. Converter os dados para CSV
    csv_file_path = 'results/driver_found.csv'  # Caminho para o arquivo CSV na pasta 'results'
    # Call the json_to_csv function to convert the data to CSV
    json_to_csv(csv_file_name="Driver_found")
    # system = wmi.WMI().Win32_OperatingSystem()[0]
    # print("Propriedades disponíveis no objeto:")
    # for prop in system.properties:
    #     print(prop)
    # Print the drivers found and the quantity of drivers
    print(drivers_found)
    print("The number of drivers is:", quantity_drivers)
